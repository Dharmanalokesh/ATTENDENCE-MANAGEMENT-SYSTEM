import cv2
from datetime import datetime, timedelta
import openpyxl
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import sqlite3
import os
import logging
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import json
from apscheduler.schedulers.background import BackgroundScheduler
from flask import Flask, render_template, request, jsonify, send_from_directory, redirect, url_for, flash, Response
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from zipfile import BadZipFile

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

app = Flask(__name__)
app.secret_key = '8f42a9b1c3d5e7f9a1b2c3d4e5f67890'

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "student_login"

EXCEL_PATH = "attendance22.xlsx"
ATTENDANCE_SHEET_PATH = "attendance_sheet.xlsx"
GOOGLE_SHEET_NAME = "Attendance_Tracker"
CREDENTIALS_PATH = "credentials.json"
CONFIG_PATH = "config.json"
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

# Email configuration for Idea #3 (replace with your email settings)
EMAIL_ADDRESS = "your-email@gmail.com"
EMAIL_PASSWORD = "your-app-password"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

# Load or initialize configuration for Idea #3
def load_config():
    default_config = {"reminders_enabled": False, "consecutive_days": 3}
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, 'r') as f:
            return json.load(f)
    else:
        with open(CONFIG_PATH, 'w') as f:
            json.dump(default_config, f)
        return default_config

def save_config(config):
    with open(CONFIG_PATH, 'w') as f:
        json.dump(config, f)

config = load_config()

class User(UserMixin):
    def __init__(self, id, role):
        self.id = id
        self.role = role

def init_db():
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
                 username TEXT PRIMARY KEY,
                 password TEXT,
                 role TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS students (
                 pin TEXT PRIMARY KEY,
                 name TEXT,
                 branch TEXT,
                 course TEXT,
                 photo_path TEXT,
                 email TEXT)''')  # Added email for Idea #3
    c.execute('''CREATE TABLE IF NOT EXISTS activity_log (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 timestamp TEXT,
                 action TEXT,
                 details TEXT)''')
    c.execute("INSERT OR IGNORE INTO users VALUES ('trainer1', 'pass456', 'trainer')")
    
    c.execute('''CREATE TABLE IF NOT EXISTS feedback (
             id INTEGER PRIMARY KEY AUTOINCREMENT,
             pin TEXT,
             comment TEXT,
             date TEXT)''')
    
    try:
        excel_data = pd.ExcelFile(EXCEL_PATH)
        for sheet_name in excel_data.sheet_names:
            df = excel_data.parse(sheet_name)
            required_columns = ["PIN (Roll.No)", "NAME", "BRANCH"]
            if not all(col in df.columns for col in required_columns):
                logging.error(f"Sheet '{sheet_name}' missing required columns: {required_columns}")
                continue
            for _, row in df.iterrows():
                pin = str(row["PIN (Roll.No)"]).strip('"')
                name = row["NAME"]
                branch = row["BRANCH"]
                c.execute("INSERT OR IGNORE INTO users VALUES (?, ?, 'student')", (pin, "LOKESH"))
                c.execute("INSERT OR IGNORE INTO students VALUES (?, ?, ?, ?, ?, ?)",
                          (pin, name, branch, sheet_name, f"static/images/{pin}.jpg", None))
    except Exception as e:
        logging.warning(f"{EXCEL_PATH} not found or corrupted during init: {e}. Skipping student import.")
    conn.commit()
    conn.close()

@login_manager.user_loader
def load_user(username):
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT username, role FROM users WHERE username = ?", (username,))
    user_data = c.fetchone()
    conn.close()
    if user_data:
        return User(user_data[0], user_data[1])
    return None

try:
    excel_data = pd.ExcelFile(EXCEL_PATH)
except Exception as e:
    logging.warning(f"Failed to load {EXCEL_PATH} at startup: {e}")
    excel_data = None

def initialize_gsheets():
    if not os.path.exists(CREDENTIALS_PATH):
        logging.error(f"{CREDENTIALS_PATH} not found.")
        return None
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_PATH, scope)
        client = gspread.authorize(creds)
        return client.open(GOOGLE_SHEET_NAME)
    except Exception as e:
        logging.error(f"Failed to initialize Google Sheets: {e}")
        return None

def sanitize_sheet_name(name, for_google_sheets=False):
    if for_google_sheets:
        return name.replace('&', '').replace('/', '_').replace(':', '').replace('*', '').replace('?', '')
    return name

def find_student_sheet_and_info(pin):
    global excel_data
    pin = str(pin).strip('"')
    if excel_data:
        try:
            for sheet_name in excel_data.sheet_names:
                df = excel_data.parse(sheet_name)
                for _, row in df.iterrows():
                    excel_pin = str(row["PIN (Roll.No)"]).strip('"')
                    if excel_pin == pin:
                        logging.info(f"Found PIN {pin} in sheet '{sheet_name}'")
                        return sheet_name, row["NAME"], row["BRANCH"]
        except Exception as e:
            logging.error(f"Error parsing {EXCEL_PATH} in find_student_sheet_and_info: {e}")
    
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT name, branch, course FROM students WHERE pin = ?", (pin,))
    student_data = c.fetchone()
    conn.close()
    if student_data:
        logging.info(f"Found PIN {pin} in database, course: {student_data[2]}")
        return student_data[2], student_data[0], student_data[1]
    
    logging.warning(f"PIN {pin} not found in Excel or database.")
    return "Unknown", f"Student_{pin}", "Unknown"

def get_excel_date_column(ws, date):
    for col in range(4, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == date:
            return col
    return None

def log_activity(action, details):
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute("INSERT INTO activity_log (timestamp, action, details) VALUES (?, ?, ?)", (timestamp, action, details))
    conn.commit()
    conn.close()

def get_recent_activity():
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT timestamp, action, details FROM activity_log ORDER BY id DESC LIMIT 5")
    activities = c.fetchall()
    conn.close()
    return activities

# Idea #3: Email Sending Function
def send_email(to_email, subject, body):
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        logging.info(f"Email sent to {to_email}: {subject}")
        return True
    except Exception as e:
        logging.error(f"Failed to send email to {to_email}: {e}")
        return False

# Idea #3: Check for Consecutive Absences and Send Reminders
def check_absent_students():
    if not config["reminders_enabled"]:
        logging.info("Attendance reminders are disabled.")
        return
    
    consecutive_days = config["consecutive_days"]
    today = datetime.now()
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT pin, name, email FROM students WHERE email IS NOT NULL")
    students = c.fetchall()
    conn.close()
    
    if not os.path.exists(ATTENDANCE_SHEET_PATH):
        logging.error(f"{ATTENDANCE_SHEET_PATH} not found.")
        return
    
    try:
        wb = openpyxl.load_workbook(ATTENDANCE_SHEET_PATH)
        for student in students:
            pin, name, email = student
            consecutive_absences = 0
            for i in range(consecutive_days):
                check_date = (today - timedelta(days=i)).strftime("%d-%m-%Y")
                found = False
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    date_col = get_excel_date_column(ws, check_date)
                    if date_col:
                        for row in range(2, ws.max_row + 1):
                            student_pin = str(ws[f"A{row}"].value).strip('"') if ws[f"A{row}"].value else ""
                            if student_pin == pin:
                                status = ws.cell(row=row, column=date_col).value
                                if status == "Absent":
                                    consecutive_absences += 1
                                found = True
                                break
                        if found:
                            break
                if not found:
                    consecutive_absences += 1
            
            if consecutive_absences >= consecutive_days:
                subject = "Attendance Reminder: You've Been Absent"
                body = f"Dear {name},\n\nYou have been absent for {consecutive_absences} consecutive days. " \
                       "Please attend classes or contact your trainer if you have any issues.\n\nBest regards,\nYour Trainer"
                if send_email(email, subject, body):
                    log_activity("Send Reminder", f"Sent absence reminder to {name} (PIN: {pin}) at {email}")
    
    except Exception as e:
        logging.error(f"Error checking absent students: {e}")

# Schedule the reminder task (Idea #3)
scheduler = BackgroundScheduler()
scheduler.add_job(check_absent_students, 'interval', days=1, start_date=datetime.now())
scheduler.start()

def update_excel(scanned_pins):
    try:
        wb = openpyxl.load_workbook(ATTENDANCE_SHEET_PATH) if os.path.exists(ATTENDANCE_SHEET_PATH) else openpyxl.Workbook()
        if not wb.sheetnames:
            wb.remove(wb.active)
    except Exception as e:
        logging.error(f"Failed to load {ATTENDANCE_SHEET_PATH}: {e}. Creating new workbook.")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    
    today_date = datetime.now().strftime("%d-%m-%Y")
    scanned_pins_set = set(scanned_pins)
    
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT pin, name, branch, course FROM students")
    db_students = {row[0]: (row[1], row[2], row[3]) for row in c.fetchall()}
    conn.close()
    
    students_by_course = {}
    for pin, (name, branch, course) in db_students.items():
        if course not in students_by_course:
            students_by_course[course] = []
        students_by_course[course].append((pin, name, branch))
    
    scanned_courses = set()
    for pin in scanned_pins_set:
        if pin in db_students:
            course = db_students[pin][2]
            if course != "Unknown":
                scanned_courses.add(course)
    
    for course in scanned_courses:
        students = students_by_course.get(course, [])
        
        if course not in wb.sheetnames:
            ws = wb.create_sheet(course)
            ws["A1"] = "PIN (Roll.No)"
            ws["B1"] = "NAME"
            ws["C1"] = "BRANCH"
            logging.info(f"Created new sheet in {ATTENDANCE_SHEET_PATH}: {course}")
        else:
            ws = wb[course]
        
        date_col = get_excel_date_column(ws, today_date)
        if not date_col:
            date_col = ws.max_column + 1 if ws.max_column >= 4 else 4
            ws.cell(row=1, column=date_col).value = today_date
        
        # Only update the current day's column, don't touch previous columns
        sheet_pin_map = {str(ws[f"A{row}"].value).strip('"'): row for row in range(2, ws.max_row + 1) if ws[f"A{row}"].value}
        for pin, name, branch in students:
            if pin not in sheet_pin_map:
                next_row = ws.max_row + 1
                ws[f"A{next_row}"] = pin
                ws[f"B{next_row}"] = name
                ws[f"C{next_row}"] = branch
                sheet_pin_map[pin] = next_row
        
        for row in range(2, ws.max_row + 1):
            pin = str(ws[f"A{row}"].value).strip('"') if ws[f"A{row}"].value else ""
            if pin:
                cell = ws.cell(row=row, column=date_col)
                # Only update the current day's status
                if pin in scanned_pins_set:
                    cell.value = "Present"
                else:
                    # Only set to "Absent" if the cell is empty for today
                    if not cell.value:
                        cell.value = "Absent"
    
    try:
        wb.save(ATTENDANCE_SHEET_PATH)
        logging.info(f"Updated {ATTENDANCE_SHEET_PATH} with {len(scanned_pins)} new scans for {today_date}")
        return True
    except Exception as e:
        logging.error(f"Failed to save {ATTENDANCE_SHEET_PATH}: {e}")
        return False

def update_google_sheets(scanned_pins):
    sheet = initialize_gsheets()
    if not sheet:
        return False
    
    today_date = datetime.now().strftime("%d-%m-%Y")
    scanned_pins_set = set(scanned_pins)
    
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT pin, name, branch, course FROM students")
    db_students = {row[0]: (row[1], row[2], row[3]) for row in c.fetchall()}
    conn.close()
    
    students_by_course = {}
    for pin, (name, branch, course) in db_students.items():
        if course not in students_by_course:
            students_by_course[course] = []
        students_by_course[course].append((pin, name, branch))
    
    scanned_courses = set()
    for pin in scanned_pins_set:
        if pin in db_students:
            course = db_students[pin][2]
            if course != "Unknown":
                scanned_courses.add(course)
    
    for course in scanned_courses:
        students = students_by_course.get(course, [])
        sheet_name = sanitize_sheet_name(course, for_google_sheets=True)
        
        try:
            ws = None
            for worksheet in sheet.worksheets():
                if worksheet.title == sheet_name:
                    ws = worksheet
                    break
            if not ws:
                ws = sheet.add_worksheet(title=sheet_name, rows=len(students) + 10, cols=20)
                ws.update('A1:C1', [['PIN (Roll.No)', 'NAME', 'BRANCH']])
            
            headers = ws.row_values(1)
            date_col = None
            for i, header in enumerate(headers, 1):
                if header == today_date:
                    date_col = i
                    break
            if not date_col:
                date_col = len(headers) + 1
                ws.update_cell(1, date_col, today_date)
            
            existing_present = set()
            all_data = ws.get_all_values()
            if len(all_data) < 2:
                all_data.append([''] * len(headers))
            
            for row in all_data[1:]:
                pin = str(row[0]).strip('"') if row else ""
                if pin and len(row) >= date_col:
                    status = row[date_col-1]
                    if status == "Present":
                        existing_present.add(pin)
            
            all_present = existing_present.union(scanned_pins_set)
            
            sheet_pins = {row[0].strip('"') for row in all_data[1:] if row and row[0]}
            updates = []
            for row_idx, row in enumerate(all_data[1:], start=2):
                pin = row[0].strip('"') if row else ""
                if not pin:
                    continue
                cell_value = row[date_col-1] if len(row) > date_col-1 else ""
                if pin in all_present:
                    if cell_value != "Present":
                        updates.append({
                            "range": f"{chr(64 + date_col)}{row_idx}",
                            "values": [["Present"]]
                        })
                else:
                    if not cell_value:
                        updates.append({
                            "range": f"{chr(64 + date_col)}{row_idx}",
                            "values": [["Absent"]]
                        })
            
            new_rows = []
            for pin, name, branch in students:
                if pin not in sheet_pins:
                    new_row = [pin, name, branch] + [""] * (date_col - 4) + ["Present" if pin in all_present else "Absent"]
                    new_rows.append(new_row)
            
            if updates:
                ws.batch_update(updates)
            if new_rows:
                ws.append_rows(new_rows)
            logging.info(f"Updated Google Sheet '{sheet_name}' with {len(scanned_pins)} new scans, total present: {len(all_present)}")
        
        except Exception as e:
            logging.error(f"Failed to update Google Sheet for course {course}: {e}")
            continue
    
    return True

# Idea #5: Update Excel and Google Sheets for Attendance Correction
def update_attendance(pin, date, new_status, course):
    # Update Excel
    try:
        wb = openpyxl.load_workbook(ATTENDANCE_SHEET_PATH)
        if course not in wb.sheetnames:
            logging.error(f"Course {course} not found in {ATTENDANCE_SHEET_PATH}")
            return False
        ws = wb[course]
        date_col = get_excel_date_column(ws, date)
        if not date_col:
            logging.error(f"Date {date} not found in sheet {course}")
            return False
        
        for row in range(2, ws.max_row + 1):
            student_pin = str(ws[f"A{row}"].value).strip('"') if ws[f"A{row}"].value else ""
            if student_pin == pin:
                ws.cell(row=row, column=date_col).value = new_status
                break
        
        wb.save(ATTENDANCE_SHEET_PATH)
    except Exception as e:
        logging.error(f"Failed to update Excel attendance for PIN {pin}: {e}")
        return False
    
    # Update Google Sheets
    sheet = initialize_gsheets()
    if not sheet:
        return False
    
    sheet_name = sanitize_sheet_name(course, for_google_sheets=True)
    try:
        ws = None
        for worksheet in sheet.worksheets():
            if worksheet.title == sheet_name:
                ws = worksheet
                break
        if not ws:
            logging.error(f"Sheet {sheet_name} not found in Google Sheets")
            return False
        
        all_data = ws.get_all_values()
        if len(all_data) < 1:
            return False
        
        headers = all_data[0]
        date_col = None
        for i, header in enumerate(headers, 1):
            if header == date:
                date_col = i
                break
        if not date_col:
            logging.error(f"Date {date} not found in Google Sheet {sheet_name}")
            return False
        
        for row_idx, row in enumerate(all_data[1:], start=2):
            student_pin = str(row[0]).strip('"') if row else ""
            if student_pin == pin:
                ws.update_cell(row_idx, date_col, new_status)
                break
        
        logging.info(f"Updated attendance for PIN {pin} on {date} to {new_status} in both Excel and Google Sheets")
        return True
    except Exception as e:
        logging.error(f"Failed to update Google Sheets attendance for PIN {pin}: {e}")
        return False

def get_excel_attendance(date):
    date_str = date.strftime("%d-%m-%Y")
    present_students = []
    absent_students = []
    
    if not os.path.exists(ATTENDANCE_SHEET_PATH):
        logging.error(f"{ATTENDANCE_SHEET_PATH} not found.")
        return [], []
    
    try:
        wb = openpyxl.load_workbook(ATTENDANCE_SHEET_PATH)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            date_col = get_excel_date_column(ws, date_str)
            if date_col:
                for row in range(2, ws.max_row + 1):
                    pin = str(ws[f"A{row}"].value).strip('"') if ws[f"A{row}"].value else ""
                    name = ws[f"B{row}"].value or "Unknown"
                    branch = ws[f"C{row}"].value or "Unknown"
                    status = ws.cell(row=row, column=date_col).value
                    if pin and status:
                        if status == "Present":
                            present_students.append((pin, name, branch, sheet_name))
                        elif status == "Absent":
                            absent_students.append((pin, name, branch, sheet_name))
        wb.close()
    except Exception as e:
        logging.error(f"Error fetching Excel attendance for {date_str}: {e}")
        return [], []
    
    return present_students, absent_students

def get_gsheets_attendance(date):
    sheet = initialize_gsheets()
    if not sheet:
        return [], []
    
    date_str = date.strftime("%d-%m-%Y")
    present_students = []
    absent_students = []
    
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT DISTINCT course FROM students")
    courses = [row[0] for row in c.fetchall()]
    conn.close()
    
    for course in courses:
        sheet_name = sanitize_sheet_name(course, for_google_sheets=True)
        try:
            ws = None
            for worksheet in sheet.worksheets():
                if worksheet.title == sheet_name:
                    ws = worksheet
                    break
            if not ws:
                continue
            
            all_data = ws.get_all_values()
            if len(all_data) < 1:
                continue
            
            headers = all_data[0]
            date_col = None
            for i, header in enumerate(headers, 1):
                if header == date_str:
                    date_col = i
                    break
            if not date_col:
                continue
            
            for row in all_data[1:]:
                if len(row) < 3:
                    continue
                pin = str(row[0]).strip('"') if row[0] else ""
                name = row[1] or "Unknown"
                branch = row[2] or "Unknown"
                status = row[date_col-1] if len(row) > date_col-1 else None
                
                if pin and status:
                    if status == "Present":
                        present_students.append((pin, name, branch))
                    elif status == "Absent":
                        absent_students.append((pin, name, branch))
        
        except Exception as e:
            logging.error(f"Error fetching Google Sheets attendance for {course}: {e}")
            continue
    
    return present_students, absent_students

def get_dashboard_stats():
    today = datetime.now()
    total_students = 0
    present_today = 0
    absent_today = 0
    missing_photos = 0
    
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM students")
    total_students = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM students WHERE photo_path IS NULL OR photo_path NOT LIKE 'static/images/%'")
    missing_photos = c.fetchone()[0]
    conn.close()
    
    present_students, absent_students = get_excel_attendance(today)
    if present_students is not None and absent_students is not None:
        present_today = len(present_students)
        absent_today = len(absent_students)
    
    percentage = (present_today / (present_today + absent_today) * 100) if (present_today + absent_today) > 0 else 0
    return total_students, present_today, absent_today, round(percentage, 2), missing_photos

@app.route('/')
def intro():
    return render_template('intro.html')

@app.route('/home')
def home():
    return render_template('home.html')

@app.route('/student_login', methods=['GET', 'POST'])
def student_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        c.execute("SELECT username, role FROM users WHERE username = ? AND password = ? AND role = 'student'", 
                  (username, password))
        user_data = c.fetchone()
        conn.close()
        if user_data:
            user = User(user_data[0], user_data[1])
            login_user(user)
            return redirect(url_for('student_dashboard'))
        flash("Invalid PIN or password")
    return render_template('student_login.html')

@app.route('/trainer_login', methods=['GET', 'POST'])
def trainer_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        c.execute("SELECT username, role FROM users WHERE username = ? AND password = ? AND role = 'trainer'", 
                  (username, password))
        user_data = c.fetchone()
        conn.close()
        if user_data:
            user = User(user_data[0], user_data[1])
            login_user(user)
            return redirect(url_for('trainer_dashboard'))
        flash("Invalid credentials")
    return render_template('trainer_login.html')

@app.route('/student_dashboard')
@login_required
def student_dashboard():
    if current_user.role != 'student':
        return redirect(url_for('trainer_dashboard'))
    
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT pin, name, branch, course, photo_path FROM students WHERE pin = ?", (current_user.id,))
    student = c.fetchone()
    conn.close()
    
    if not student:
        return "Student data not found", 404
    
    wb = openpyxl.load_workbook(ATTENDANCE_SHEET_PATH) if os.path.exists(ATTENDANCE_SHEET_PATH) else openpyxl.Workbook()
    if student[3] in wb.sheetnames:
        ws = wb[student[3]]
        date_cols = [(col, ws.cell(row=1, column=col).value) for col in range(4, ws.max_column + 1) if ws.cell(row=1, column=col).value]
        total_days = len(date_cols)
        
        present_days = 0
        absent_days = 0
        recent_history = []
        for col, date in date_cols:
            for row in range(2, ws.max_row + 1):
                if str(ws[f"A{row}"].value).strip('"') == student[0]:
                    status = ws.cell(row=row, column=col).value
                    if status == "Present":
                        present_days += 1
                    elif status == "Absent":
                        absent_days += 1
                    recent_history.append((date, status if status in ["Present", "Absent"] else "Not Marked"))
                    break
        
        recent_history.sort(key=lambda x: datetime.strptime(x[0], "%d-%m-%Y"), reverse=True)
        recent_history = recent_history[:5]
        
        percentage = (present_days / total_days * 100) if total_days > 0 else 0
        low_attendance = percentage < 75
    else:
        total_days = 0
        present_days = 0
        absent_days = 0
        recent_history = []
        percentage = 0
        low_attendance = False
    
    return render_template('student_dashboard.html', student=student, percentage=round(percentage, 2), 
                           present_days=present_days, absent_days=absent_days, total_days=total_days,
                           recent_history=recent_history, low_attendance=low_attendance)

@app.route('/submit_feedback', methods=['POST'])
@login_required
def submit_feedback():
    if current_user.role != 'student':
        flash("Only students can submit feedback.", "error")
        return redirect(url_for('trainer_dashboard'))
    try:
        comment = request.form['comment']
        date = datetime.now().strftime('%Y-%m-%d')
        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        c.execute("INSERT INTO feedback (pin, comment, date) VALUES (?, ?, ?)",
                  (current_user.id, comment, date))
        c.execute("INSERT INTO activity_log (timestamp, action, details) VALUES (?, ?, ?)",
                  (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Feedback Submitted", f"Student {current_user.id} submitted feedback"))
        conn.commit()
        flash("Feedback submitted successfully!", "success")
    except Exception as e:
        flash(f"Error submitting feedback: {str(e)}", "error")
        logging.error(f"Error in submit_feedback: {str(e)}")
    finally:
        conn.close()
    return redirect(url_for('student_dashboard'))

# Idea #3: Add Toggle for Reminders
@app.route('/toggle_reminders', methods=['POST'])
@login_required
def toggle_reminders():
    if current_user.role != 'trainer':
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    
    enabled = request.form.get('enabled') == 'true'
    config["reminders_enabled"] = enabled
    save_config(config)
    log_activity("Toggle Reminders", f"Reminders {'enabled' if enabled else 'disabled'}")
    return jsonify({"status": "success", "message": f"Reminders {'enabled' if enabled else 'disabled'}"})

@app.route('/review_feedback')
@login_required
def review_feedback():
    if current_user.role != 'trainer':
        flash("Only trainers can view feedback.", "error")
        return redirect(url_for('student_dashboard'))
    try:
        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        import os
        logging.info(f"Connected to database: {os.path.abspath('database.db')}")
        c.execute("SELECT pin, comment, date FROM feedback ORDER BY date DESC")
        feedbacks = c.fetchall()
        logging.info(f"Retrieved feedbacks before render: {feedbacks}")
        if not feedbacks:
            logging.info("No feedback records found in the database.")
        else:
            logging.info(f"Number of feedback records: {len(feedbacks)}")
    except Exception as e:
        flash(f"Error retrieving feedback: {str(e)}", "error")
        logging.error(f"Error in review_feedback: {str(e)}")
        feedbacks = []
    finally:
        conn.close()
    return render_template('trainer_dashboard.html', action='review_feedback', feedbacks=feedbacks)

@app.route('/trainer_dashboard', methods=['GET', 'POST'])
@login_required
def trainer_dashboard():
    if current_user.role != 'trainer':
        return redirect(url_for('student_dashboard'))
    
    action = request.args.get('action', None)
    total_students, present_today, absent_today, percentage_today, missing_photos = get_dashboard_stats()
    recent_activity = get_recent_activity()
    default_date = datetime.now().strftime('%Y-%m-%d')
    reminders_enabled = config["reminders_enabled"]  # Idea #3: Pass to template
    
    if action == 'search':
        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        search_query = request.args.get('search', '').strip()
        search_field = request.args.get('search_field', 'pin')
        field_map = {'pin': 'pin', 'name': 'name', 'branch': 'branch', 'course': 'course'}
        column = field_map.get(search_field, 'pin')
        
        if search_query:
            c.execute(f"SELECT pin, name, branch, course, photo_path FROM students WHERE {column} LIKE ? ORDER BY pin",
                      (f'%{search_query}%',))
        else:
            c.execute("SELECT pin, name, branch, course, photo_path FROM students ORDER BY pin")
        
        students = c.fetchall()
        conn.close()
        return render_template('trainer_dashboard.html', action=action, students=students,
                              total_students=total_students, present_today=present_today, 
                              absent_today=absent_today, percentage_today=percentage_today, 
                              missing_photos=missing_photos, recent_activity=recent_activity,
                              default_date=default_date, reminders_enabled=reminders_enabled)
    
    elif action == 'today_excel':
        date_str = request.args.get('date', default_date)
        try:
            date = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            date = datetime.now()
            flash("Invalid date format. Showing today’s data.")
        
        selected_branch = request.args.get('branch', 'all')
        selected_course = request.args.get('course', 'all')
        
        present_students, absent_students = get_excel_attendance(date)
        if present_students is None or absent_students is None:
            flash("Error fetching Excel attendance data. Ensure the Excel file exists and is accessible.")
            present_students, absent_students = [], []
        
        if selected_branch != 'all':
            present_students = [student for student in present_students if student[2] == selected_branch]
            absent_students = [student for student in absent_students if student[2] == selected_branch]
        
        if selected_course != 'all':
            present_students = [student for student in present_students if student[3] == selected_course]
            absent_students = [student for student in absent_students if student[3] == selected_course]
        
        branches = ['all', 'ECE', 'CSE', 'CSC', 'EEE', 'CSM', 'CSD']
        courses = ['all', 'AWS&JAVA', 'GENAI', 'REDHAT&AWS', 'JFS&UIUX']
        
        return render_template('trainer_dashboard.html', action=action, 
                              present_students=present_students, absent_students=absent_students,
                              selected_date=date.strftime("%d-%m-%Y"),
                              total_students=total_students, present_today=present_today, 
                              absent_today=absent_today, percentage_today=percentage_today, 
                              missing_photos=missing_photos, recent_activity=recent_activity,
                              default_date=default_date, branches=branches, courses=courses,
                              selected_branch=selected_branch, selected_course=selected_course,
                              reminders_enabled=reminders_enabled)
    
    elif action == 'today_gsheets':
        date_str = request.args.get('date', default_date)
        try:
            date = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            date = datetime.now()
            flash("Invalid date format. Showing today’s data.")
        present_students, absent_students = get_gsheets_attendance(date)
        if present_students is None or absent_students is None:
            flash("Error fetching Google Sheets attendance data. Check credentials and connectivity.")
        return render_template('trainer_dashboard.html', action=action, 
                              present_students=present_students, absent_students=absent_students,
                              selected_date=date.strftime("%d-%m-%Y"),
                              total_students=total_students, present_today=present_today, 
                              absent_today=absent_today, percentage_today=percentage_today, 
                              missing_photos=missing_photos, recent_activity=recent_activity,
                              default_date=default_date, reminders_enabled=reminders_enabled)
    
    return render_template('trainer_dashboard.html', action=action,
                          total_students=total_students, present_today=present_today, 
                          absent_today=absent_today, percentage_today=percentage_today, 
                          missing_photos=missing_photos, recent_activity=recent_activity,
                          default_date=default_date, reminders_enabled=reminders_enabled)

@app.route('/scan', methods=['POST'])
@login_required
def scan():
    if current_user.role != 'trainer':
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    
    data = request.get_json()
    scanned_pins = data.get('scanned_pins', [])
    
    if scanned_pins:
        try:
            update_excel(scanned_pins)
            update_google_sheets(scanned_pins)
            for pin in scanned_pins:
                log_activity("Scan QR", f"Scanned PIN {pin}")
            logging.info(f"Scanned PINs: {scanned_pins}")
            return jsonify({"status": "success", "scanned": scanned_pins})
        except Exception as e:
            logging.error(f"Failed to process scan: {e}")
            return jsonify({"status": "error", "message": f"Error processing scan: {str(e)}"}), 500
    return jsonify({"status": "error", "message": "No QR codes scanned"})

@app.route('/add_student', methods=['POST'])
@login_required
def add_student():
    if current_user.role != 'trainer':
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    
    pin = request.form['pin']
    name = request.form['name']
    branch = request.form['branch']
    course = request.form['course']
    email = request.form['email']  # Idea #3: Added email field
    
    if 'photo' not in request.files:
        return jsonify({"status": "error", "message": "No photo uploaded"}), 400
    photo = request.files['photo']
    if photo.filename == '':
        return jsonify({"status": "error", "message": "No photo selected"}), 400
    
    photo_path = f"static/images/{pin}.jpg"
    os.makedirs("static/images", exist_ok=True)
    try:
        photo.save(photo_path)
    except Exception as e:
        logging.error(f"Failed to save photo: {e}")
        return jsonify({"status": "error", "message": f"Failed to save photo: {e}"}), 500
    
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO students (pin, name, branch, course, photo_path, email) VALUES (?, ?, ?, ?, ?, ?)",
              (pin, name, branch, course, photo_path, email))
    c.execute("INSERT OR IGNORE INTO users VALUES (?, ?, 'student')", (pin, "LOKESH"))
    conn.commit()
    conn.close()
    
    wb = openpyxl.load_workbook(EXCEL_PATH) if os.path.exists(EXCEL_PATH) else openpyxl.Workbook()
    sheet_name = course
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = "PIN (Roll.No)"
        ws["B1"] = "NAME"
        ws["C1"] = "BRANCH"
    ws = wb[sheet_name]
    sheet_pin_map = {str(ws[f"A{row}"].value).strip('"'): row for row in range(2, ws.max_row + 1) if ws[f"A{row}"].value}
    if pin not in sheet_pin_map:
        next_row = ws.max_row + 1
        ws[f"A{next_row}"] = pin
        ws[f"B{next_row}"] = name
        ws[f"C{next_row}"] = branch
    wb.save(EXCEL_PATH)
    
    log_activity("Add Student", f"Added student {name} (PIN: {pin})")
    return redirect(url_for('trainer_dashboard'))

# Idea #8: Bulk Upload for Adding Students
@app.route('/bulk_upload_students', methods=['POST'])
@login_required
def bulk_upload_students():
    if current_user.role != 'trainer':
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    
    if 'file' not in request.files:
        flash("No file uploaded")
        return redirect(url_for('trainer_dashboard', action='add'))
    
    file = request.files['file']
    if file.filename == '':
        flash("No file selected")
        return redirect(url_for('trainer_dashboard', action='add'))
    
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        flash("Invalid file format. Please upload a CSV or Excel file.")
        return redirect(url_for('trainer_dashboard', action='add'))
    
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        required_columns = ["PIN (Roll.No)", "NAME", "BRANCH", "COURSE", "EMAIL"]
        if not all(col in df.columns for col in required_columns):
            flash(f"File must contain columns: {', '.join(required_columns)}")
            return redirect(url_for('trainer_dashboard', action='add'))
        
        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        added_count = 0
        
        wb = openpyxl.load_workbook(EXCEL_PATH) if os.path.exists(EXCEL_PATH) else openpyxl.Workbook()
        
        for _, row in df.iterrows():
            pin = str(row["PIN (Roll.No)"]).strip('"')
            name = row["NAME"]
            branch = row["BRANCH"]
            course = row["COURSE"]
            email = row["EMAIL"]
            
            # Skip if email is missing or invalid (basic validation)
            if not isinstance(email, str) or '@' not in email:
                continue
            
            # Add to database
            c.execute("INSERT OR REPLACE INTO students (pin, name, branch, course, photo_path, email) VALUES (?, ?, ?, ?, ?, ?)",
                      (pin, name, branch, course, None, email))
            c.execute("INSERT OR IGNORE INTO users VALUES (?, ?, 'student')", (pin, "LOKESH"))
            
            # Add to attendance22.xlsx
            sheet_name = course
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws["A1"] = "PIN (Roll.No)"
                ws["B1"] = "NAME"
                ws["C1"] = "BRANCH"
            ws = wb[sheet_name]
            sheet_pin_map = {str(ws[f"A{row}"].value).strip('"'): row for row in range(2, ws.max_row + 1) if ws[f"A{row}"].value}
            if pin not in sheet_pin_map:
                next_row = ws.max_row + 1
                ws[f"A{next_row}"] = pin
                ws[f"B{next_row}"] = name
                ws[f"C{next_row}"] = branch
                added_count += 1
        
        wb.save(EXCEL_PATH)
        conn.commit()
        conn.close()
        
        log_activity("Bulk Upload", f"Added {added_count} students via bulk upload")
        flash(f"Successfully added {added_count} students")
    except Exception as e:
        flash(f"Error processing file: {str(e)}")
        logging.error(f"Error in bulk upload: {e}")
    
    return redirect(url_for('trainer_dashboard', action='add'))

# Idea #5: Attendance Correction Route
@app.route('/correct_attendance', methods=['POST'])
@login_required
def correct_attendance():
    if current_user.role != 'trainer':
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    
    data = request.get_json()
    pin = data.get('pin')
    date = data.get('date')
    new_status = data.get('status')
    course = data.get('course')  # Course is now passed from the frontend
    
    if not all([pin, date, new_status, course]):
        return jsonify({"status": "error", "message": "Missing required fields"}), 400
    
    if new_status not in ["Present", "Absent"]:
        return jsonify({"status": "error", "message": "Invalid status"}), 400
    
    if update_attendance(pin, date, new_status, course):
        log_activity("Correct Attendance", f"Changed attendance for PIN {pin} on {date} to {new_status} in course {course}")
        return jsonify({"status": "success", "message": f"Attendance updated to {new_status}"})
    else:
        return jsonify({"status": "error", "message": "Failed to update attendance"}), 500

@app.route('/download_excel_today')
@login_required
def download_excel_today():
    date_str = request.args.get('date', datetime.now().strftime("%d-%m-%Y"))
    try:
        date = datetime.strptime(date_str, "%d-%m-%Y")
    except ValueError:
        date = datetime.now()
    
    present_students, absent_students = get_excel_attendance(date)
    if present_students is None or absent_students is None:
        return "Error generating Excel file", 500
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance_{date_str}"
    
    headers = ["PIN", "Name", "Branch", "Course", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    row = 2
    for student in present_students:
        ws.cell(row=row, column=1, value=student[0])
        ws.cell(row=row, column=2, value=student[1])
        ws.cell(row=row, column=3, value=student[2])
        ws.cell(row=row, column=4, value=student[3])
        ws.cell(row=row, column=5, value="Present")
        row += 1
    for student in absent_students:
        ws.cell(row=row, column=1, value=student[0])
        ws.cell(row=row, column=2, value=student[1])
        ws.cell(row=row, column=3, value=student[2])
        ws.cell(row=row, column=4, value=student[3])
        ws.cell(row=row, column=5, value="Absent")
        row += 1
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_activity("Download Excel", f"Downloaded Excel for {date_str}")
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename=excel_attendance_{date_str}.xlsx"})

@app.route('/download_gsheets_today')
@login_required
def download_gsheets_today():
    date_str = request.args.get('date', datetime.now().strftime("%d-%m-%Y"))
    try:
        date = datetime.strptime(date_str, "%d-%m-%Y")
    except ValueError:
        date = datetime.now()
    
    present_students, absent_students = get_gsheets_attendance(date)
    if present_students is None or absent_students is None:
        return "Error generating Excel file", 500
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance_{date_str}"
    
    headers = ["PIN", "Name", "Branch", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    row = 2
    for student in present_students:
        ws.cell(row=row, column=1, value=student[0])
        ws.cell(row=row, column=2, value=student[1])
        ws.cell(row=row, column=3, value=student[2])
        ws.cell(row=row, column=4, value="Present")
        row += 1
    for student in absent_students:
        ws.cell(row=row, column=1, value=student[0])
        ws.cell(row=row, column=2, value=student[1])
        ws.cell(row=row, column=3, value=student[2])
        ws.cell(row=row, column=4, value="Absent")
        row += 1
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_activity("Download GSheets", f"Downloaded GSheets for {date_str}")
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename=gsheets_attendance_{date_str}.xlsx"})

@app.route('/download_excel_presentees')
@login_required
def download_excel_presentees():
    if current_user.role != 'trainer':
        flash("Only trainers can access this feature.", "error")
        return redirect(url_for('trainer_dashboard'))
    date_str = request.args.get('date', datetime.now().strftime("%d-%m-%Y"))
    try:
        date = datetime.strptime(date_str, "%d-%m-%Y")
    except ValueError:
        date = datetime.now()
        flash("Invalid date format. Showing today’s data.")
    
    present_students, _ = get_excel_attendance(date)
    if present_students is None:
        return "Error generating Excel file", 500
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Presentees_{date_str}"
    
    headers = ["PIN", "Name", "Branch", "Course", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    row = 2
    for student in present_students:
        ws.cell(row=row, column=1, value=student[0])
        ws.cell(row=row, column=2, value=student[1])
        ws.cell(row=row, column=3, value=student[2])
        ws.cell(row=row, column=4, value=student[3])
        ws.cell(row=row, column=5, value="Present")
        row += 1
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_activity("Download Excel Presentees", f"Downloaded Excel for presentees on {date_str}")
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename=presentees_{date_str}.xlsx"})

@app.route('/download_excel_absentees')
@login_required
def download_excel_absentees():
    if current_user.role != 'trainer':
        flash("Only trainers can access this feature.", "error")
        return redirect(url_for('trainer_dashboard'))
    date_str = request.args.get('date', datetime.now().strftime("%d-%m-%Y"))
    try:
        date = datetime.strptime(date_str, "%d-%m-%Y")
    except ValueError:
        date = datetime.now()
        flash("Invalid date format. Showing today’s data.")
    
    _, absent_students = get_excel_attendance(date)
    if absent_students is None:
        return "Error generating Excel file", 500
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Absentees_{date_str}"
    
    headers = ["PIN", "Name", "Branch", "Course", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    row = 2
    for student in absent_students:
        ws.cell(row=row, column=1, value=student[0])
        ws.cell(row=row, column=2, value=student[1])
        ws.cell(row=row, column=3, value=student[2])
        ws.cell(row=row, column=4, value=student[3])
        ws.cell(row=row, column=5, value="Absent")
        row += 1
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_activity("Download Excel Absentees", f"Downloaded Excel for absentees on {date_str}")
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename=absentees_{date_str}.xlsx"})

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('home'))

@app.route('/static/images/<filename>')
def serve_image(filename):
    return send_from_directory('static/images', filename)

@app.route('/static/videos/<filename>')
def serve_video(filename):
    return send_from_directory('static/videos', filename)

@app.route('/static/sounds/<filename>')
def serve_sound(filename):
    return send_from_directory('static/sounds', filename)

if __name__ == "__main__":
    init_db()
    app.run(debug=True)
