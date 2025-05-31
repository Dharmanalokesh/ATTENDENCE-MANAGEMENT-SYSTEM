import cv2
from pyzbar.pyzbar import decode
from datetime import datetime
import openpyxl
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import sys
import os
import logging
import time
# Setup logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Predefined courses (Excel subsheets)
COURSES = ["AWS&JAVA", "GENAI", "JFS&UIUX", "REDHAT&AWS","JFS"]  # Updated "REDHAT&AWS" to match your log

# File and Google Sheets setup
EXCEL_PATH = "attendance22.xlsx"
GOOGLE_SHEET_NAME = "Attendance_Tracker"
CREDENTIALS_PATH = "credentials.json"
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

# Initialize Google Sheets client
def initialize_gsheets():
    if not os.path.exists(CREDENTIALS_PATH):
        logging.error(f"{CREDENTIALS_PATH} not found. Please add the Service Account JSON file.")
        return None
    
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_PATH, scope)
        client = gspread.authorize(creds)
        sheet = client.open(GOOGLE_SHEET_NAME)
        logging.info("Google Sheets connected successfully.")
        return sheet
    except Exception as e:
        logging.error(f"Error initializing Google Sheets: {e}")
        return None

# Load Excel data
try:
    excel_data = pd.ExcelFile(EXCEL_PATH)
except FileNotFoundError:
    logging.warning(f"{EXCEL_PATH} not found. Initializing new file.")
    excel_data = None

# Function to scan QR codes with error handling
def scan_qr_code():
    cap = cv2.VideoCapture(0)
    scanned_pins = set()
    logging.info("Scanning QR codes... Press 'q' to stop.")
    
    while True:
        ret, frame = cap.read()
        if not ret:
            logging.error("Failed to grab frame.")
            break
        
        try:
            decoded_objects = decode(frame)
            for obj in decoded_objects:
                pin = obj.data.decode("utf-8").strip('"')  # Remove quotes
                if pin not in scanned_pins:
                    scanned_pins.add(pin)
                    logging.info(f"Scanned: {pin}")
        except Exception as e:
            logging.warning(f"Error decoding QR frame: {e}")
            continue
        
        cv2.imshow("QR Scanner", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    
    cap.release()
    cv2.destroyAllWindows()
    return scanned_pins

# Function to find student’s course sheet and info in Excel
def find_student_sheet_and_info(pin):
    if not excel_data:
        return None, None, None
    
    pin = str(pin).strip('"')  # Normalize PIN
    for sheet_name in excel_data.sheet_names:
        df = excel_data.parse(sheet_name)
        for index, row in df.iterrows():
            excel_pin = str(row["PIN (Roll.No)"]).strip('"')
            if excel_pin == pin:
                return sheet_name, row["NAME"], row["BRANCH"]
    return None, f"Student_{pin}", "Unknown"

# Function to get or set date column in Excel
def get_excel_date_column(ws, today_date):
    for col in range(4, ws.max_column + 1):  # Start from D (4th column)
        if ws.cell(row=1, column=col).value == today_date:
            return col
    # If date not found, use next available column
    date_col = ws.max_column + 1 if ws.max_column >= 4 else 4
    ws.cell(row=1, column=date_col, value=today_date)
    return date_col

# Function to update Excel with error handling
def update_excel(scanned_pins):
    try:
        if os.path.exists(EXCEL_PATH):
            wb = openpyxl.load_workbook(EXCEL_PATH)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
    except Exception as e:
        logging.error(f"Failed to load {EXCEL_PATH}: {e}. Creating new workbook.")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    
    today_date = datetime.now().strftime("%d-%m-%Y")
    pin_map = {}
    
    for pin in scanned_pins:
        sheet_name, name, branch = find_student_sheet_and_info(pin)
        if not sheet_name:
            logging.warning(f"Student {pin} not found in any sheet. Skipping.")
            continue
        
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws["A1"] = "PIN (Roll.No)"
            ws["B1"] = "NAME"
            ws["C1"] = "BRANCH"
        else:
            ws = wb[sheet_name]
        
        sheet_pin_map = {}
        for row in range(2, ws.max_row + 1):
            p = str(ws[f"A{row}"].value).strip('"')
            if p:
                sheet_pin_map[p] = (ws[f"B{row}"].value, ws[f"C{row}"].value)
        
        if pin not in sheet_pin_map:
            next_row = ws.max_row + 1
            ws[f"A{next_row}"] = pin
            ws[f"B{next_row}"] = name
            ws[f"C{next_row}"] = branch
            sheet_pin_map[pin] = (name, branch)
        
        date_col = get_excel_date_column(ws, today_date)
        for row in range(2, ws.max_row + 1):
            p = str(ws[f"A{row}"].value).strip('"')
            # Only update if cell is empty or it’s a scanned PIN
            if not ws.cell(row=row, column=date_col).value or p in scanned_pins:
                ws.cell(row=row, column=date_col, value="Present" if p in scanned_pins else "Absent")
        
        pin_map[pin] = (name, branch, sheet_name)  # Include course
        logging.info(f"Attendance for {pin} updated in {sheet_name} sheet under {today_date}")
    
    try:
        wb.save(EXCEL_PATH)
    except Exception as e:
        logging.error(f"Failed to save {EXCEL_PATH}: {e}. Saving to backup file.")
        wb.save("attendance_backup.xlsx")
    return pin_map

# Function to get or set date column in Google Sheets
def get_gsheets_date_column(worksheet, today_date):
    header_row = worksheet.row_values(1)
    for col_idx, header in enumerate(header_row, 1):
        if header == today_date:
            return col_idx
    # If date not found, append to next column
    date_col = len(header_row) + 1 if len(header_row) >= 3 else 3
    worksheet.update_cell(1, date_col, today_date)
    return date_col

# Function to update Google Sheets
def update_google_sheets(sheet, scanned_pins, pin_map):
    if not sheet:
        logging.warning("Google Sheets not available.")
        return
    
    today_date = datetime.now().strftime("%d-%m-%Y")
    try:
        branch_sheets = [ws.title for ws in sheet.worksheets()]
        logging.info(f"Found Google Sheets subsheets: {branch_sheets}")
    except Exception as e:
        logging.error(f"Error fetching subsheets: {e}")
        branch_sheets = ["ECE", "EEE", "CSD", "CSM", "CSC", "CSE","IT"]  # Fallback
    
    for branch in branch_sheets:
        try:
            worksheet = sheet.worksheet(branch)
            records = worksheet.get_all_values()
            logging.info(f"Processing {branch} subsheet with {len(records)} rows.")
            
            if not records or len(records[0]) < 2:
                worksheet.update_cell(1, 1, "PIN (Roll.No)")
                worksheet.update_cell(1, 2, "COURSE")
                records = [["PIN (Roll.No)", "COURSE"]]
                logging.info(f"Initialized {branch} subsheet headers.")
            
            # Get or set date column
            date_col = get_gsheets_date_column(worksheet, today_date)
            
            # Batch update all rows to Absent for today’s column only if not already set
            if len(records) > 1:
                col_letter = chr(64 + date_col)  # e.g., 3 -> "C"
                current_values = worksheet.col_values(date_col)[1:]  # Skip header
                if not current_values or len(current_values) < len(records) - 1:
                    absent_values = [["Absent"] for _ in range(len(records) - 1)]
                    worksheet.update(f"{col_letter}2:{col_letter}{len(records)}", absent_values)
                    logging.info(f"Batch marked {len(records) - 1} rows as Absent in {branch} subsheet for {today_date}.")
            
            # Update scanned students to Present
            for pin in pin_map:
                pin = str(pin).strip('"')  # Normalize PIN
                name, student_branch, course = pin_map[pin]
                logging.info(f"Checking PIN {pin} with branch {student_branch} against {branch}.")
                if student_branch.upper() == branch.upper():  # Case-insensitive match
                    pin_found = False
                    for row in range(2, len(records) + 1):
                        sheet_pin = str(records[row-1][0]).strip('"')
                        logging.info(f"Comparing {sheet_pin} with {pin}")
                        if sheet_pin == pin:
                            worksheet.update_cell(row, date_col, "Present")
                            logging.info(f"Updated {branch} sheet: {pin} - Present at row {row} for {today_date}")
                            pin_found = True
                            break
                    if not pin_found and pin in scanned_pins:
                        worksheet.append_row([pin, course] + [""] * (date_col - 3) + ["Present"])
                        logging.info(f"Appended {branch} sheet: {pin} - Present with course {course} for {today_date}")
            time.sleep(1)  # Avoid quota limit
        except Exception as e:
            logging.error(f"Error processing {branch} subsheet: {e}")
            time.sleep(5)  # Longer delay on error

# CLI to select course
def select_course():
    print("\nAvailable Courses (for reference):")
    for i, course in enumerate(COURSES, 1):
        print(f"{i}. {course}")
    print(f"{len(COURSES) + 1}. Exit")
    
    while True:
        try:
            choice = int(input("Select a course or exit (enter number): "))
            if 1 <= choice <= len(COURSES):
                return COURSES[choice - 1]
            elif choice == len(COURSES) + 1:
                return None
            else:
                print("⚠️ Invalid choice. Try again.")
        except ValueError:
            print("⚠️ Please enter a valid number.")

# Main execution
def main():
    sheet = initialize_gsheets()
    all_scanned_pins = set()
    
    while True:
        course = select_course()
        if not course:
            print("⏹️ Exiting program.")
            break
        
        scanned_pins = scan_qr_code()
        if scanned_pins:
            all_scanned_pins.update(scanned_pins)
            pin_map = update_excel(scanned_pins)
            update_google_sheets(sheet, scanned_pins, pin_map)
        else:
            logging.warning("No QR codes scanned.")
        
        while True:
            cont = input("Continue scanning? (y/n): ").lower()
            if cont in ['y', 'n']:
                break
            print("⚠️ Please enter 'y' or 'n'.")
        if cont != 'y':
            print("⏹️ Exiting program.")
            break

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        sys.exit(1)